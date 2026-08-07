import uuid
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response
import io
import csv
from sqlalchemy.orm import Session

from database import get_db, SessionLocal
from Auth.models import User
from Auth.router import get_current_user
from FileUpload.BlobFile import upload_file
import module.FestivalDB as FestivalDB
from Festival import common
from Festival import EmailSending

router = APIRouter(prefix="/festivals", tags=["Festival Wishes"])

# Automated startup migration for columns added after the table already
# existed in production (Base.metadata.create_all only creates new tables,
# it never alters existing ones).
try:
    from database import engine
    from sqlalchemy import text
    with engine.connect() as conn:
        for col in [
            "ALTER TABLE festival_wishes ADD COLUMN audience VARCHAR DEFAULT 'employees';",
            "ALTER TABLE festival_wishes ADD COLUMN cc_emails VARCHAR;",
            "ALTER TABLE festival_wishes ADD COLUMN from_email VARCHAR;",
            "ALTER TABLE festival_wishes ADD COLUMN template_id INTEGER;",
            "ALTER TABLE wish_templates ADD COLUMN logo_url VARCHAR;",
            "ALTER TABLE wish_templates ADD COLUMN logo_width INTEGER DEFAULT 120;",
            "ALTER TABLE wish_templates ADD COLUMN logo_align VARCHAR DEFAULT 'center';",
            "ALTER TABLE wish_templates ADD COLUMN company_name VARCHAR;",
            "ALTER TABLE wish_templates ADD COLUMN company_website VARCHAR;",
            "ALTER TABLE wish_templates ADD COLUMN company_email VARCHAR;",
            "ALTER TABLE wish_templates ADD COLUMN company_phone VARCHAR;",
            "ALTER TABLE wish_templates ADD COLUMN company_tagline VARCHAR;",
            "ALTER TABLE festival_wishes ADD COLUMN to_emails VARCHAR;",
            "ALTER TABLE commercial_emails ADD COLUMN to_emails VARCHAR;",
            "ALTER TABLE festival_wishes ADD COLUMN send_time VARCHAR DEFAULT '09:00';",
            "ALTER TABLE email_provider_config ADD COLUMN batch_size INTEGER DEFAULT 30;",
            "ALTER TABLE email_provider_config ADD COLUMN delay_seconds INTEGER DEFAULT 0;",
            "ALTER TABLE wish_contacts ADD COLUMN company_name VARCHAR;",
            "ALTER TABLE wish_contacts ADD COLUMN contact_type VARCHAR DEFAULT 'customer';",
            "ALTER TABLE commercial_emails ADD COLUMN no_template BOOLEAN DEFAULT FALSE;",
        ]:
            try:
                conn.execute(text(col))
                conn.commit()
            except Exception:
                conn.rollback()
except Exception:
    pass


def _is_today(wish: FestivalDB.FestivalWish, today: date) -> bool:
    if wish.recurs_yearly:
        return wish.date.month == today.month and wish.date.day == today.day
    return wish.date == today


def _serialize(w: FestivalDB.FestivalWish):
    return {
        "id": w.id,
        "name": w.name,
        "date": w.date.isoformat(),
        "send_time": w.send_time or "09:00",
        "message": w.message,
        "recurs_yearly": w.recurs_yearly,
        "enabled": w.enabled,
        "audience": w.audience or "employees",
        "to_emails": w.to_emails or "",
        "cc_emails": w.cc_emails or "",
        "from_email": w.from_email or "",
        "template_id": w.template_id,
    }


def _serialize_template(t: FestivalDB.WishTemplate):
    return {
        "id": t.id,
        "name": t.name,
        "company_name": t.company_name or "",
        "company_website": t.company_website or "",
        "company_email": t.company_email or "",
        "company_phone": t.company_phone or "",
        "company_tagline": t.company_tagline or "",
        "header_html": t.header_html,
        "header_bg_color": t.header_bg_color,
        "highlight_html": t.highlight_html or "",
        "highlight_bg_color": t.highlight_bg_color,
        "footer_html": t.footer_html,
        "footer_bg_color": t.footer_bg_color,
        "is_default": t.is_default,
        "logo_url": t.logo_url or "",
        "logo_width": t.logo_width or 120,
        "logo_align": t.logo_align or "center",
    }


def _serialize_log(log: FestivalDB.WishSendLog):
    return {
        "id": log.id,
        "recipient_name": log.recipient_name,
        "to_email": log.to_email,
        "cc_emails": log.cc_emails,
        "from_email": log.from_email,
        "status": log.status,
        "error": log.error,
        "sent_at": log.sent_at.isoformat() if log.sent_at else None,
    }


def _serialize_contact(c: FestivalDB.WishContact):
    return {
        "id": c.id,
        "name": c.name,
        "email": c.email,
        "company_name": c.company_name or "",
        "contact_type": c.contact_type or "customer",
        "enabled": c.enabled,
    }


@router.get("/today")
def get_today_wish(db: Session = Depends(get_db)):
    today = date.today()
    wishes = db.query(FestivalDB.FestivalWish).filter(FestivalDB.FestivalWish.enabled == True).all()
    for w in wishes:
        if _is_today(w, today):
            return _serialize(w)
    return None


@router.get("")
def list_wishes(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    wishes = db.query(FestivalDB.FestivalWish).order_by(FestivalDB.FestivalWish.date).all()
    return [_serialize(w) for w in wishes]


@router.post("")
def create_wish(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    try:
        wish_date = datetime.strptime(data["date"], "%Y-%m-%d").date()
    except (KeyError, ValueError):
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    audience = data.get("audience", "employees")
    if audience not in ("employees", "customers", "both"):
        audience = "employees"
    wish = FestivalDB.FestivalWish(
        name=data.get("name", "").strip() or "Festival",
        date=wish_date,
        send_time=data.get("send_time", "09:00").strip() or "09:00",
        message=data.get("message", "").strip() or "Wishing you a happy celebration!",
        recurs_yearly=bool(data.get("recurs_yearly", True)),
        enabled=bool(data.get("enabled", True)),
        audience=audience,
        to_emails=data.get("to_emails", "").strip() or None,
        cc_emails=data.get("cc_emails", "").strip() or None,
        from_email=data.get("from_email", "").strip() or None,
        template_id=data.get("template_id") or None,
    )
    db.add(wish)
    db.commit()
    db.refresh(wish)
    return _serialize(wish)


@router.put("/{wish_id}")
def update_wish(wish_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    wish = db.query(FestivalDB.FestivalWish).filter(FestivalDB.FestivalWish.id == wish_id).first()
    if not wish:
        raise HTTPException(status_code=404, detail="Not found")
    if "name" in data:
        wish.name = data["name"].strip() or wish.name
    if "date" in data:
        try:
            wish.date = datetime.strptime(data["date"], "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    if "send_time" in data:
        wish.send_time = data["send_time"].strip() or "09:00"
    if "message" in data:
        wish.message = data["message"].strip() or wish.message
    if "recurs_yearly" in data:
        wish.recurs_yearly = bool(data["recurs_yearly"])
    if "enabled" in data:
        wish.enabled = bool(data["enabled"])
    if "audience" in data and data["audience"] in ("employees", "customers", "both"):
        wish.audience = data["audience"]
    if "to_emails" in data:
        wish.to_emails = data["to_emails"].strip() or None
    if "cc_emails" in data:
        wish.cc_emails = data["cc_emails"].strip() or None
    if "from_email" in data:
        wish.from_email = data["from_email"].strip() or None
    if "template_id" in data:
        wish.template_id = data["template_id"] or None
    db.commit()
    db.refresh(wish)
    return _serialize(wish)


@router.delete("/{wish_id}")
def delete_wish(wish_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    wish = db.query(FestivalDB.FestivalWish).filter(FestivalDB.FestivalWish.id == wish_id).first()
    if not wish:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(wish)
    db.commit()
    return {"message": "Deleted"}


# ── Email templates (header/footer/colors) — reusable across festivals ──

@router.get("/templates")
def list_templates(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    templates = db.query(FestivalDB.WishTemplate).order_by(FestivalDB.WishTemplate.name).all()
    return [_serialize_template(t) for t in templates]


@router.post("/templates")
def create_template(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    if not data.get("name", "").strip():
        raise HTTPException(status_code=400, detail="name is required")
    if data.get("is_default"):
        db.query(FestivalDB.WishTemplate).update({"is_default": False})
    template = FestivalDB.WishTemplate(
        name=data["name"].strip(),
        company_name=data.get("company_name", "").strip() or None,
        company_website=data.get("company_website", "").strip() or None,
        company_email=data.get("company_email", "").strip() or None,
        company_phone=data.get("company_phone", "").strip() or None,
        company_tagline=data.get("company_tagline", "").strip() or None,
        header_html=data.get("header_html", "").strip() or "<div>{{festival_name}}</div>",
        header_bg_color=data.get("header_bg_color") or "#9EE4FF",
        highlight_html=data.get("highlight_html", "").strip() or None,
        highlight_bg_color=data.get("highlight_bg_color") or "#9EE4FF",
        footer_html=data.get("footer_html", "").strip() or "<div>Contact us</div>",
        footer_bg_color=data.get("footer_bg_color") or "#FAFBFD",
        is_default=bool(data.get("is_default", False)),
        logo_url=data.get("logo_url", "").strip() or None,
        logo_width=data.get("logo_width") or 120,
        logo_align=data.get("logo_align") or "center",
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return _serialize_template(template)


@router.put("/templates/{template_id}")
def update_template(template_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    template = db.query(FestivalDB.WishTemplate).filter(FestivalDB.WishTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Not found")
    if "name" in data:
        template.name = data["name"].strip() or template.name
    if "company_name" in data:
        template.company_name = data["company_name"].strip() or None
    if "company_website" in data:
        template.company_website = data["company_website"].strip() or None
    if "company_email" in data:
        template.company_email = data["company_email"].strip() or None
    if "company_phone" in data:
        template.company_phone = data["company_phone"].strip() or None
    if "company_tagline" in data:
        template.company_tagline = data["company_tagline"].strip() or None
    if "header_html" in data:
        template.header_html = data["header_html"]
    if "header_bg_color" in data:
        template.header_bg_color = data["header_bg_color"] or template.header_bg_color
    if "highlight_html" in data:
        template.highlight_html = data["highlight_html"] or None
    if "highlight_bg_color" in data:
        template.highlight_bg_color = data["highlight_bg_color"] or template.highlight_bg_color
    if "footer_html" in data:
        template.footer_html = data["footer_html"]
    if "footer_bg_color" in data:
        template.footer_bg_color = data["footer_bg_color"] or template.footer_bg_color
    if "logo_url" in data:
        template.logo_url = data["logo_url"].strip() or None
    if "logo_width" in data:
        template.logo_width = data["logo_width"] or template.logo_width
    if "logo_align" in data:
        template.logo_align = data["logo_align"] or template.logo_align
    if data.get("is_default"):
        db.query(FestivalDB.WishTemplate).filter(FestivalDB.WishTemplate.id != template_id).update({"is_default": False})
        template.is_default = True
    elif "is_default" in data:
        template.is_default = bool(data["is_default"])
    db.commit()
    db.refresh(template)
    return _serialize_template(template)


@router.delete("/templates/{template_id}")
def delete_template(template_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    template = db.query(FestivalDB.WishTemplate).filter(FestivalDB.WishTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Not found")
    if template.is_default:
        raise HTTPException(status_code=400, detail="Cannot delete the default template — set another one as default first")
    db.query(FestivalDB.FestivalWish).filter(FestivalDB.FestivalWish.template_id == template_id).update({"template_id": None})
    db.delete(template)
    db.commit()
    return {"message": "Deleted"}


# ── Customer / external contacts (audience for "customers" / "both" sends) ──

@router.get("/contacts")
def list_contacts(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    contacts = db.query(FestivalDB.WishContact).order_by(FestivalDB.WishContact.name).all()
    return [_serialize_contact(c) for c in contacts]


@router.post("/contacts")
def create_contact(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    name = data.get("name", "").strip()
    email = data.get("email", "").strip()
    company_name = data.get("company_name", "").strip() or None
    contact_type = data.get("contact_type", "customer").strip().lower()
    if contact_type not in ("customer", "employee", "both"):
        contact_type = "customer"
    if not name or not email:
        raise HTTPException(status_code=400, detail="name and email are required")
    contact = FestivalDB.WishContact(
        name=name,
        email=email,
        company_name=company_name,
        contact_type=contact_type,
        enabled=bool(data.get("enabled", True))
    )
    db.add(contact)
    db.commit()
    db.refresh(contact)
    return _serialize_contact(contact)


@router.put("/contacts/{contact_id}")
def update_contact(contact_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    contact = db.query(FestivalDB.WishContact).filter(FestivalDB.WishContact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Not found")
    if "name" in data:
        contact.name = data["name"].strip() or contact.name
    if "email" in data:
        contact.email = data["email"].strip() or contact.email
    if "company_name" in data:
        contact.company_name = data["company_name"].strip() or None
    if "contact_type" in data:
        ctype = str(data["contact_type"]).strip().lower()
        if ctype in ("customer", "employee", "both"):
            contact.contact_type = ctype
    if "enabled" in data:
        contact.enabled = bool(data["enabled"])
    db.commit()
    db.refresh(contact)
    return _serialize_contact(contact)


@router.delete("/contacts/{contact_id}")
def delete_contact(contact_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    contact = db.query(FestivalDB.WishContact).filter(FestivalDB.WishContact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(contact)
    db.commit()
    return {"message": "Deleted"}


@router.post("/contacts/sync-employees")
def sync_employees_from_sso(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Pull SSO-synced login accounts from the users table into the audience
    contact list, tagged as 'employee'. Idempotent: emails that already exist
    as contacts (enabled or disabled) are skipped, so a disabled contact is
    never re-added."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    existing = {
        c.email.strip().lower()
        for c in db.query(FestivalDB.WishContact.email).all()
        if c.email and c.email.strip()
    }

    users = db.query(User).filter(User.email.isnot(None), User.email != "").all()
    added = 0
    for u in users:
        email = (u.email or "").strip()
        if not email or "@" not in email:
            continue
        if email.lower() in existing:
            continue
        name = ((getattr(u, "username", None) or "").strip()) or email.split("@")[0]
        db.add(FestivalDB.WishContact(
            name=name,
            email=email,
            company_name=None,
            contact_type="employee",
            enabled=True,
        ))
        existing.add(email.lower())
        added += 1

    db.commit()
    return {
        "message": f"Synced {added} new employee contact(s) from SSO users."
        if added else "All SSO employees are already in the contact list.",
        "added": added,
        "total_users": len(users),
    }


@router.get("/contacts/sample-excel")
def download_sample_excel():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Company Name", "Contact Name", "Contact Email", "Contact Type"])
    writer.writerow(["TIBOS Solutions", "John Doe", "john@tibos.in", "customer"])
    writer.writerow(["Acme Corporation", "Jane Smith", "jane@acme.com", "employee"])
    writer.writerow(["Global Tech", "Alex Lee", "alex@global.com", "both"])
    content = output.getvalue()
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="audience_contacts_sample.csv"'}
    )


@router.post("/contacts/bulk-upload")
async def bulk_upload_contacts(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    filename = (file.filename or "").lower()
    content = await file.read()

    rows = []
    if filename.endswith(".csv") or filename.endswith(".txt"):
        text_data = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text_data))
        for r in reader:
            rows.append(r)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for h, val in zip(headers, row):
                    if h:
                        row_dict[h] = str(val or "").strip()
                rows.append(row_dict)
        except Exception:
            text_data = content.decode("utf-8-sig", errors="ignore")
            reader = csv.DictReader(io.StringIO(text_data))
            for r in reader:
                rows.append(r)
    else:
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are supported")

    added_count = 0
    updated_count = 0
    skipped_count = 0

    for r in rows:
        norm_r = {str(k).strip().lower().replace(" ", "_"): str(v or "").strip() for k, v in r.items() if k}
        company_name = norm_r.get("company_name") or norm_r.get("company") or norm_r.get("organization") or None
        name = norm_r.get("contact_name") or norm_r.get("customer_name") or norm_r.get("name") or norm_r.get("client_name") or ""
        email = norm_r.get("contact_email") or norm_r.get("customer_email") or norm_r.get("email") or norm_r.get("email_id") or norm_r.get("mail") or ""
        contact_type = norm_r.get("contact_type") or norm_r.get("audience_type") or norm_r.get("audience") or norm_r.get("type") or "customer"
        contact_type = contact_type.strip().lower()
        if contact_type not in ("customer", "employee", "both"):
            contact_type = "customer"

        if not email or "@" not in email:
            skipped_count += 1
            continue

        if not name:
            name = email.split("@")[0].capitalize()

        existing = db.query(FestivalDB.WishContact).filter(FestivalDB.WishContact.email == email).first()
        if existing:
            existing.name = name
            existing.company_name = company_name or existing.company_name
            existing.contact_type = contact_type or existing.contact_type or "customer"
            existing.enabled = True
            updated_count += 1
        else:
            new_contact = FestivalDB.WishContact(
                name=name,
                email=email,
                company_name=company_name,
                contact_type=contact_type,
                enabled=True
            )
            db.add(new_contact)
            added_count += 1

    db.commit()
    return {
        "message": f"Bulk upload completed: {added_count} contacts added, {updated_count} updated.",
        "added": added_count,
        "updated": updated_count,
        "skipped": skipped_count
    }


@router.get("/all-send-history")
def get_all_send_history(
    status: str = None,
    campaign_type: str = None,
    search: str = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    f_logs = db.query(FestivalDB.WishSendLog).order_by(FestivalDB.WishSendLog.sent_at.desc()).limit(1000).all()
    c_logs = db.query(FestivalDB.CommercialSendLog).order_by(FestivalDB.CommercialSendLog.sent_at.desc()).limit(1000).all()

    combined = []
    for l in f_logs:
        combined.append({
            "id": f"f_{l.id}",
            "campaign_type": "Festival Wish",
            "campaign_name": l.wish_name or "Festival Wish",
            "recipient_name": l.recipient_name or "",
            "to_email": l.to_email,
            "from_email": l.from_email or "",
            "cc_emails": l.cc_emails or "",
            "status": l.status,
            "error": l.error or None,
            "sent_at": l.sent_at.isoformat() if l.sent_at else "",
        })

    for l in c_logs:
        combined.append({
            "id": f"c_{l.id}",
            "campaign_type": "Commercial Email",
            "campaign_name": l.email_name or "Commercial Email",
            "recipient_name": l.recipient_name or "",
            "to_email": l.to_email,
            "from_email": l.from_email or "",
            "cc_emails": l.cc_emails or "",
            "status": l.status,
            "error": l.error or None,
            "sent_at": l.sent_at.isoformat() if l.sent_at else "",
        })

    combined.sort(key=lambda x: x["sent_at"] or "", reverse=True)

    if search and search.strip():
        q = search.strip().lower()
        combined = [
            x for x in combined
            if q in x["to_email"].lower()
            or q in x["recipient_name"].lower()
            or q in x["campaign_name"].lower()
            or q in (x["error"] or "").lower()
        ]

    if status and status.strip() and status != "all":
        st = status.strip().lower()
        combined = [x for x in combined if x["status"] == st]

    if campaign_type and campaign_type.strip() and campaign_type != "all":
        ct = campaign_type.strip().lower()
        combined = [x for x in combined if ct in x["campaign_type"].lower()]

    total_triggered = len(combined)
    delivered_count = sum(1 for x in combined if x["status"] == "sent")
    failed_count = sum(1 for x in combined if x["status"] == "failed")
    success_rate = round((delivered_count / total_triggered * 100)) if total_triggered > 0 else 100

    return {
        "summary": {
            "total": total_triggered,
            "delivered": delivered_count,
            "failed": failed_count,
            "success_rate": success_rate,
        },
        "logs": combined
    }


@router.get("/delivery-dashboard-stats")
def get_delivery_dashboard_stats(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    from sqlalchemy import or_
    import module.EmplyeeDB as EmplyeeDB

    f_logs = db.query(FestivalDB.WishSendLog).order_by(FestivalDB.WishSendLog.sent_at.desc()).limit(1000).all()
    c_logs = db.query(FestivalDB.CommercialSendLog).order_by(FestivalDB.CommercialSendLog.sent_at.desc()).limit(1000).all()

    total_triggered = len(f_logs) + len(c_logs)
    delivered_f = sum(1 for l in f_logs if l.status == "sent")
    delivered_c = sum(1 for l in c_logs if l.status == "sent")
    total_delivered = delivered_f + delivered_c
    failed_f = sum(1 for l in f_logs if l.status == "failed")
    failed_c = sum(1 for l in c_logs if l.status == "failed")
    total_failed = failed_f + failed_c

    success_rate = round((total_delivered / total_triggered * 100)) if total_triggered > 0 else 100

    total_contacts = db.query(FestivalDB.WishContact).count()
    emp_contacts = db.query(FestivalDB.WishContact).filter(FestivalDB.WishContact.contact_type == "employee").count()
    cust_contacts = db.query(FestivalDB.WishContact).filter(
        or_(FestivalDB.WishContact.contact_type == "customer", FestivalDB.WishContact.contact_type.is_(None))
    ).count()
    both_contacts = db.query(FestivalDB.WishContact).filter(FestivalDB.WishContact.contact_type == "both").count()

    active_employees = db.query(EmplyeeDB.Employee).filter(
        EmplyeeDB.Employee.email.isnot(None),
        EmplyeeDB.Employee.email != ""
    ).count()

    return {
        "summary": {
            "total": total_triggered,
            "delivered": total_delivered,
            "failed": total_failed,
            "success_rate": success_rate,
        },
        "campaign_breakdown": {
            "festival_wishes": {"total": len(f_logs), "delivered": delivered_f, "failed": failed_f},
            "commercial_emails": {"total": len(c_logs), "delivered": delivered_c, "failed": failed_c},
        },
        "audience_stats": {
            "system_employees": active_employees,
            "customer_contacts": cust_contacts,
            "employee_contacts": emp_contacts,
            "both_contacts": both_contacts,
            "total_custom_contacts": total_contacts,
        }
    }


@router.post("/upload-image")
def upload_wish_image(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    url = upload_file(file.file, file.filename or f"{uuid.uuid4()}.png", folder="festival_images")
    return {"url": url}


# ── Email sending configuration (Microsoft 365 / Google) — shared by ──
# ── both Festival Wishes and Commercial Emails ──

@router.get("/email-config")
def get_email_config(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    cfg = EmailSending.get_or_create_config(db)
    return {
        "provider": cfg.provider or "",
        "batch_size": cfg.batch_size if cfg.batch_size is not None else 30,
        "delay_seconds": cfg.delay_seconds if cfg.delay_seconds is not None else 0,
        "microsoft": {
            "client_id": cfg.ms_client_id or "",
            "client_secret_set": bool(cfg.ms_client_secret),
            "tenant": cfg.ms_tenant or "common",
            "sender_email": cfg.ms_sender_email or "",
        },
        "google": {
            "service_account_set": bool(cfg.google_service_account_json),
            "sender_email": cfg.google_sender_email or "",
        },
    }


@router.post("/email-config")
def save_email_config(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    provider = data.get("provider")
    if provider not in ("microsoft", "google"):
        raise HTTPException(status_code=400, detail="provider must be 'microsoft' or 'google'")
    cfg = EmailSending.get_or_create_config(db)
    cfg.provider = provider
    if "batch_size" in data:
        try:
            cfg.batch_size = max(1, int(data["batch_size"]))
        except (ValueError, TypeError):
            pass
    if "delay_seconds" in data:
        try:
            cfg.delay_seconds = max(0, int(data["delay_seconds"]))
        except (ValueError, TypeError):
            pass

    ms = data.get("microsoft", {})
    if ms.get("client_id"):
        cfg.ms_client_id = ms["client_id"].strip()
    if ms.get("client_secret"):
        cfg.ms_client_secret = ms["client_secret"].strip()
    if ms.get("tenant"):
        cfg.ms_tenant = ms["tenant"].strip()
    if "sender_email" in ms:
        cfg.ms_sender_email = ms["sender_email"].strip() or None

    google = data.get("google", {})
    if google.get("service_account_json"):
        cfg.google_service_account_json = google["service_account_json"].strip()
    if "sender_email" in google:
        cfg.google_sender_email = google["sender_email"].strip() or None

    db.commit()
    return {"message": "Email configuration saved"}


@router.post("/email-config/test")
async def test_email_config(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    to_email = data.get("to_email", "").strip()
    if not to_email:
        raise HTTPException(status_code=400, detail="to_email is required")
    ok, err = await EmailSending.send_email(
        db,
        "TIBOS HRMS — Test Email",
        "<p>This is a test email from your HRMS Celebrations email configuration. If you received this, sending is working correctly.</p>",
        to_email,
    )
    if not ok:
        raise HTTPException(status_code=502, detail=err or "Test email failed to send")
    return {"message": f"Test email sent to {to_email}"}


@router.get("/{wish_id}/logs")
def get_wish_logs(wish_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    logs = (
        db.query(FestivalDB.WishSendLog)
        .filter(FestivalDB.WishSendLog.wish_id == wish_id)
        .order_by(FestivalDB.WishSendLog.sent_at.desc())
        .all()
    )
    return [_serialize_log(l) for l in logs]


async def send_wish_email(wish: FestivalDB.FestivalWish, db: Session):
    recipients = common.get_recipients(wish.audience, db, wish.cc_emails or "", wish.to_emails or "")
    if not recipients:
        return False, "No recipient emails found. Please add active employees in Employee Management, or add customer emails under Customer Contacts."

    template = common.get_template(wish.template_id, db)
    if not template:
        return False, "No email template configured — add one from Celebrations > Email Templates"

    cfg = EmailSending.get_or_create_config(db)
    batch_size = cfg.batch_size or 30
    delay_seconds = cfg.delay_seconds or 0

    sender_override = (wish.from_email or "").strip() or None
    cc_list = [c.strip() for c in (wish.cc_emails or "").split(",") if c.strip()]
    subject = f"🎉 {wish.name} Wishes from TIBOS"
    sent, failed = 0, 0
    total = len(recipients)

    for idx, (name, email) in enumerate(recipients):
        body_html = common.build_email_html(wish.name, template, common.merge_message(wish.message, name))
        ok, err = await EmailSending.send_email(db, subject, body_html, email, cc_list, sender_override)
        db.add(FestivalDB.WishSendLog(
            wish_id=wish.id,
            wish_name=wish.name,
            recipient_name=name,
            to_email=email,
            cc_emails=wish.cc_emails,
            from_email=sender_override,
            status="sent" if ok else "failed",
            error=None if ok else err,
        ))
        if ok:
            sent += 1
        else:
            failed += 1

        # Batch delay: pause if batch_size reached and more recipients remain
        if (idx + 1) % batch_size == 0 and (idx + 1) < total and delay_seconds > 0:
            await asyncio.sleep(delay_seconds)

    db.commit()

    if sent == 0:
        return False, f"All {failed} email(s) failed to send"
    return True, f"Sent to {sent} recipient(s)" + (f", {failed} failed" if failed else "")


@router.post("/send-now/{wish_id}")
async def send_now(wish_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    wish = db.query(FestivalDB.FestivalWish).filter(FestivalDB.FestivalWish.id == wish_id).first()
    if not wish:
        raise HTTPException(status_code=404, detail="Not found")
    ok, detail = await send_wish_email(wish, db)
    if not ok:
        raise HTTPException(status_code=502, detail=f"Email not sent: {detail}")
    wish.last_email_sent_year = date.today().year
    db.commit()
    return {"message": detail}


async def run_daily_festival_check():
    """Called periodically by scheduler. Evaluates both festival date AND
    scheduled send_time before triggering automatic wishes."""
    today = date.today()
    now_time_str = datetime.now().strftime("%H:%M")
    db = SessionLocal()
    try:
        wishes = db.query(FestivalDB.FestivalWish).filter(FestivalDB.FestivalWish.enabled == True).all()
        for w in wishes:
            if not _is_today(w, today):
                continue
            wish_time = (w.send_time or "09:00").strip()
            if now_time_str >= wish_time:
                claimed = db.query(FestivalDB.FestivalWish).filter(
                    FestivalDB.FestivalWish.id == w.id,
                    (FestivalDB.FestivalWish.last_email_sent_year.is_(None))
                    | (FestivalDB.FestivalWish.last_email_sent_year != today.year),
                ).update({"last_email_sent_year": today.year}, synchronize_session=False)
                db.commit()
                if claimed:
                    ok, detail = await send_wish_email(w, db)
                    if not ok:
                        print(f"[festival] email for '{w.name}' not sent: {detail}")
    finally:
        db.close()


def seed_default_festivals():
    """Seeds a handful of fixed-date national holidays. Festivals that move
    every year (Diwali, Holi, Eid, Pongal, etc.) are intentionally left out —
    add them with the correct date for the year from Admin > Festival Wishes."""
    db = SessionLocal()
    try:
        if db.query(FestivalDB.FestivalWish).count() > 0:
            return
        defaults = [
            ("New Year", date(2026, 1, 1), "Wishing you a very Happy New Year! May this year bring success and happiness to you and your family."),
            ("Republic Day", date(2026, 1, 26), "Happy Republic Day! Celebrating the spirit of our nation together."),
            ("Independence Day", date(2026, 8, 15), "Happy Independence Day! Proud to celebrate this day with our team."),
            ("Gandhi Jayanti", date(2026, 10, 2), "Happy Gandhi Jayanti! Remembering the values of truth and non-violence."),
            ("Christmas", date(2026, 12, 25), "Merry Christmas! Wishing you and your family a joyful holiday season."),
        ]
        for name, d, msg in defaults:
            db.add(FestivalDB.FestivalWish(name=name, date=d, message=msg, recurs_yearly=True, enabled=True))
        db.commit()
    finally:
        db.close()


def seed_default_template():
    """Seeds one editable default template matching the original hardcoded
    design, so existing wishes keep looking the same until an admin
    customizes it from Celebrations > Email Templates."""
    db = SessionLocal()
    try:
        if db.query(FestivalDB.WishTemplate).count() > 0:
            return
        db.add(FestivalDB.WishTemplate(
            name="Default TIBOS Theme",
            header_html=(
                '<div style="line-height:120%;margin:0 0 8pt;font-family:Verdana, Geneva, sans-serif;font-size:18pt;font-weight:bold;">{{festival_name}}</div>'
                '<div style="line-height:120%;margin:0 0 8pt;font-family:\'Trebuchet MS\', Trebuchet, sans-serif;font-size:20pt;font-weight:bold;">🎉 Wishing You a Very Happy {{festival_name}}! 🎉</div>'
            ),
            header_bg_color="#9EE4FF",
            highlight_html=(
                '<div style="line-height:1.38;margin:0 0 8pt;font-size:11pt;font-weight:bold;">From all of us at TIBOS 💛</div>'
                '<div style="line-height:1.38;font-size:11pt;">Wishing you and your family joy, prosperity and togetherness.</div>'
            ),
            highlight_bg_color="#9EE4FF",
            footer_html=(
                '<div style="line-height:1.38;margin:0 0 8pt;font-size:11pt;font-weight:bold;color:#000;">TIBOS Solutions &amp; Services Pvt. Ltd.</div>'
                '<div style="line-height:1.38;margin:0 0 8pt;font-size:11pt;">'
                '🌐 <a href="http://www.tibos.co.in" style="color:#467886;">www.tibos.co.in</a>'
                '&nbsp;&nbsp;&nbsp;|&nbsp;&nbsp;&nbsp;📧 <a href="mailto:secure@tibos.in" style="color:#467886;">secure@tibos.in</a>'
                '&nbsp;&nbsp;&nbsp;|&nbsp;&nbsp;📞 +91 92821 09750'
                '</div>'
                '<div style="line-height:1.38;font-size:11pt;color:#000;">Wishing our entire team a joyful and safe celebration.</div>'
            ),
            footer_bg_color="#FAFBFD",
            is_default=True,
        ))
        db.commit()
    finally:
        db.close()
